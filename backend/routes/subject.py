from flask import Blueprint, request, jsonify, make_response
from models import db, Subject, QuestionPaper, AnswerSheet, EvaluationRubric, Mark, User
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

subject_bp = Blueprint('subject', __name__)

@subject_bp.route('', methods=['POST'])
def create_subject():
    """Create a new subject/class.
    
    Accepts optional first_evaluator_id and second_evaluator_id to assign
    evaluators at creation time (custodian workflow).
    """
    try:
        data = request.json
        name = data.get('name')
        class_name = data.get('className')
        academic_year = data.get('academicYear')
        first_evaluator_id = data.get('first_evaluator_id')
        second_evaluator_id = data.get('second_evaluator_id')
        created_by = data.get('created_by')  # Optional: custodian user id
        
        if not name:
            return jsonify({'error': 'Subject name is required'}), 400
        
        # Validate evaluator IDs if provided
        if first_evaluator_id:
            if not User.query.get(first_evaluator_id):
                return jsonify({'error': 'First evaluator not found'}), 404
        if second_evaluator_id:
            if not User.query.get(second_evaluator_id):
                return jsonify({'error': 'Second evaluator not found'}), 404
        
        subject = Subject(
            name=name,
            class_name=class_name,
            academic_year=academic_year,
            first_evaluator_id=first_evaluator_id,
            second_evaluator_id=second_evaluator_id,
            created_by=created_by
        )
        
        db.session.add(subject)
        db.session.commit()
        
        print(f"✅ Created subject: {name} (ID={subject.id})")
        
        return jsonify({
            'message': 'Subject created successfully',
            'subject': subject.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error creating subject: {str(e)}")
        return jsonify({'error': str(e)}), 500


@subject_bp.route('', methods=['GET'])
def list_subjects():
    """Get all subjects"""
    try:
        subjects = Subject.query.order_by(Subject.created_at.desc()).all()
        
        return jsonify({
            'subjects': [s.to_dict() for s in subjects]
        }), 200
        
    except Exception as e:
        print(f"❌ Error listing subjects: {str(e)}")
        return jsonify({'error': str(e)}), 500


@subject_bp.route('/<int:subject_id>', methods=['GET'])
def get_subject(subject_id):
    """Get subject details with associated files"""
    try:
        subject = Subject.query.get_or_404(subject_id)
        
        # Get associated files
        question_papers = QuestionPaper.query.filter_by(subject_id=subject_id).all()
        rubrics = EvaluationRubric.query.filter_by(subject_id=subject_id).all()
        answer_sheets = AnswerSheet.query.filter_by(subject_id=subject_id).all()
        
        return jsonify({
            'subject': subject.to_dict(),
            'question_papers': [qp.to_dict() for qp in question_papers],
            'rubrics': [r.to_dict() for r in rubrics],
            'answer_sheets': [a.to_dict() for a in answer_sheets],
            'stats': {
                'total_students': len(answer_sheets),
                'evaluated': len([a for a in answer_sheets if a.status == 'evaluated']),
                'pending': len([a for a in answer_sheets if a.status != 'evaluated'])
            }
        }), 200
        
    except Exception as e:
        print(f"❌ Error getting subject: {str(e)}")
        return jsonify({'error': str(e)}), 500


@subject_bp.route('/<int:subject_id>', methods=['DELETE'])
def delete_subject(subject_id):
    """Delete a subject and all associated files"""
    try:
        subject = Subject.query.get_or_404(subject_id)
        
        # Cascade delete will handle all relationships
        db.session.delete(subject)
        db.session.commit()
        
        print(f"✅ Deleted subject: {subject.name} (ID={subject_id})")
        
        return jsonify({
            'message': 'Subject deleted successfully'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error deleting subject: {str(e)}")
        return jsonify({'error': str(e)}), 500


@subject_bp.route('/<int:subject_id>', methods=['PUT'])
def update_subject(subject_id):
    """Update subject name, class name, and academic year."""
    try:
        subject = Subject.query.get_or_404(subject_id)
        data = request.json

        name = (data.get('name') or '').strip()
        if not name:
            return jsonify({'error': 'Subject name is required'}), 400

        subject.name = name
        if 'className' in data:
            subject.class_name = (data['className'] or '').strip() or None
        if 'academicYear' in data:
            subject.academic_year = (data['academicYear'] or '').strip() or None

        db.session.commit()
        print(f"✅ Updated subject {subject_id}: {subject.name}")

        return jsonify({
            'message': 'Subject updated successfully',
            'subject': subject.to_dict()
        }), 200

    except Exception as e:
        db.session.rollback()
        print(f"❌ Error updating subject: {str(e)}")
        return jsonify({'error': str(e)}), 500


@subject_bp.route('/<int:subject_id>/assign-evaluators', methods=['PUT'])
def assign_evaluators(subject_id):
    """Assign or reassign first and second evaluators to a subject.
    
    Body: { first_evaluator_id: int|null, second_evaluator_id: int|null }
    Custodian-only operation (no JWT enforcement here to keep backward compat;
    the frontend enforces this via role-based routing).
    """
    try:
        subject = Subject.query.get_or_404(subject_id)
        data = request.json

        first_id = data.get('first_evaluator_id')
        second_id = data.get('second_evaluator_id')

        # Validate evaluator IDs
        if first_id is not None:
            if not User.query.get(first_id):
                return jsonify({'error': 'First evaluator not found'}), 404
            subject.first_evaluator_id = first_id

        if second_id is not None:
            if not User.query.get(second_id):
                return jsonify({'error': 'Second evaluator not found'}), 404
            subject.second_evaluator_id = second_id

        # Allow explicit null to unassign
        if 'first_evaluator_id' in data and data['first_evaluator_id'] is None:
            subject.first_evaluator_id = None
        if 'second_evaluator_id' in data and data['second_evaluator_id'] is None:
            subject.second_evaluator_id = None

        db.session.commit()

        print(f"✅ Evaluators assigned for subject {subject_id}: "
              f"first={subject.first_evaluator_id}, second={subject.second_evaluator_id}")

        return jsonify({
            'message': 'Evaluators assigned successfully',
            'subject': subject.to_dict()
        }), 200

    except Exception as e:
        db.session.rollback()
        print(f"❌ Error assigning evaluators: {str(e)}")
        return jsonify({'error': str(e)}), 500


@subject_bp.route('/<int:subject_id>/students', methods=['GET'])
def get_subject_students(subject_id):
    """Get all students for a subject (for student switcher)"""
    try:
        subject = Subject.query.get_or_404(subject_id)
        answer_sheets = AnswerSheet.query.filter_by(subject_id=subject_id).order_by(AnswerSheet.roll_number).all()
        
        students = []
        for sheet in answer_sheets:
            students.append({
                'id': sheet.id,
                'name': sheet.student_name,
                'roll_number': sheet.roll_number,
                'status': sheet.status,
                'file_path': sheet.file_path
            })
        
        return jsonify({
            'subject_id': subject_id,
            'subject_name': subject.name,
            'students': students
        }), 200
        
    except Exception as e:
        print(f"❌ Error getting students: {str(e)}")
        return jsonify({'error': str(e)}), 500
@subject_bp.route('/<int:subject_id>/export-marks', methods=['GET'])
def export_subject_marks(subject_id):
    """Export all marks for a subject to Excel"""
    try:
        subject = Subject.query.get_or_404(subject_id)
        answer_sheets = AnswerSheet.query.filter_by(subject_id=subject_id).order_by(AnswerSheet.roll_number).all()
        
        # Create a mapping of student marks
        # student_id -> {question_number -> marks_awarded}
        all_marks = {}
        unique_questions = set()
        
        for sheet in answer_sheets:
            marks = Mark.query.filter_by(answer_sheet_id=sheet.id).all()
            student_marks = {}
            for m in marks:
                student_marks[m.question_number] = m.marks_awarded
                unique_questions.add(m.question_number)
            all_marks[sheet.id] = student_marks
        
        # Sort unique questions (numeric if possible, else string)
        def sort_key(q):
            try:
                # Try to extract numbers if it's like "Q1", "2a"
                import re
                nums = re.findall(r'\d+', str(q))
                return int(nums[0]) if nums else 0
            except:
                return 0

        sorted_questions = sorted(list(unique_questions), key=sort_key)
        
        # Create Excel Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Marks - {subject.name}"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal="center")
        
        # Set Headers
        headers = ['Roll Number', 'Student Name'] + \
                  [f"Q{q}" if not str(q).startswith('Q') else q for q in sorted_questions] + \
                  ['Teacher Marks', 'External Marks', 'Final Average', 'Status']
        
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header_title
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            
            # Adjust column width
            column_letter = openpyxl.utils.get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = max(len(str(header_title)) + 2, 12)
            
        # Write Data
        for row_num, sheet in enumerate(answer_sheets, 2):
            ws.cell(row=row_num, column=1, value=sheet.roll_number)
            ws.cell(row=row_num, column=2, value=sheet.student_name)
            
            row_total = 0
            max_total = 0 # This would ideally come from the rubric/QP
            
            for col_num, q_num in enumerate(sorted_questions, 3):
                val = all_marks.get(sheet.id, {}).get(q_num, '-')
                ws.cell(row=row_num, column=col_num, value=val)
                if isinstance(val, (int, float)):
                    row_total += val
            
            # For percentage, we need max marks. If not available per student, we use evaluated percentage if stored.
            ws.cell(row=row_num, column=len(headers) - 3, value=sheet.teacher_marks if sheet.teacher_marks is not None else '-')
            ws.cell(row=row_num, column=len(headers) - 2, value=sheet.external_marks if sheet.external_marks is not None else '-')
            ws.cell(row=row_num, column=len(headers) - 1, value=sheet.final_marks if sheet.final_marks is not None else '-')
            ws.cell(row=row_num, column=len(headers), value=sheet.status)
            
        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{subject.name}_{subject.class_name}_Marks.xlsx".replace(" ", "_")
        
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        print(f"✅ Exported marks for subject ID={subject_id}")
        return response
        
    except Exception as e:
        print(f"❌ Error exporting marks: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def get_total_marks_logic(answer_sheet_id):
    """Helper for total marks calculation.
    Computes total max marks from unique questions.
    For awarded marks, uses the AnswerSheet.final_marks if available, 
    otherwise falls back to teacher_marks or external_marks if only one exists,
    or manually computes average if neither are yet finalized on the sheet.
    """
    try:
        from models import AnswerSheet, Mark
        sheet = AnswerSheet.query.get(answer_sheet_id)
        if not sheet: return {'total_awarded': 0, 'total_max': 0, 'percentage': 0}
        
        marks = Mark.query.filter_by(answer_sheet_id=answer_sheet_id).all()
        if not marks: return {'total_awarded': 0, 'total_max': 0, 'percentage': 0}
        
        # Calculate Max Marks (sum of unique questions)
        unique_questions = {}
        for m in marks:
            if m.question_number not in unique_questions:
                unique_questions[m.question_number] = m.max_marks
        total_max = sum(unique_questions.values())
        
        # Calculate Awarded Marks
        if sheet.final_marks is not None:
            total_awarded = sheet.final_marks
        elif sheet.teacher_marks is not None:
            total_awarded = sheet.teacher_marks
        elif sheet.external_marks is not None:
            total_awarded = sheet.external_marks
        else:
            # Fallback (e.g. legacy or partial)
            total_awarded = 0
            for q_num in unique_questions:
                q_marks = [m.marks_awarded for m in marks if m.question_number == q_num]
                if q_marks:
                    total_awarded += sum(q_marks) / len(q_marks)
                
        return {
            'total_awarded': total_awarded,
            'total_max': total_max,
            'percentage': (total_awarded / total_max * 100) if total_max > 0 else 0
        }
    except Exception as e:
        print(f"Error in get_total_marks_logic: {str(e)}")
        return None


@subject_bp.route('/<int:subject_id>/results', methods=['GET'])
def get_subject_results(subject_id):
    """Get all student results for a subject (for results library)"""
    try:
        subject = Subject.query.get_or_404(subject_id)
        answer_sheets = AnswerSheet.query.filter_by(subject_id=subject_id).order_by(AnswerSheet.roll_number).all()

        all_questions = set()
        students = []

        for sheet in answer_sheets:
            marks = Mark.query.filter_by(answer_sheet_id=sheet.id).all()
            marks_map = {}
            for m in marks:
                marks_map[str(m.question_number)] = {
                    'awarded': m.marks_awarded,
                    'max': m.max_marks
                }
                all_questions.add(str(m.question_number))

            totals = get_total_marks_logic(sheet.id)

            students.append({
                'id': sheet.id,
                'name': sheet.student_name or 'Unknown',
                'roll_number': sheet.roll_number or '-',
                'status': sheet.status,
                'marks': marks_map,
                'total_awarded': totals['total_awarded'] if totals else 0,
                'total_max': totals['total_max'] if totals else 0,
                'percentage': round(totals['percentage'], 1) if totals else 0
            })

        # Sort questions naturally
        import re
        def sort_key(q):
            nums = re.findall(r'\d+', str(q))
            return int(nums[0]) if nums else 0
        sorted_questions = sorted(list(all_questions), key=sort_key)

        return jsonify({
            'success': True,
            'subject': subject.to_dict(),
            'questions': sorted_questions,
            'students': students,
            'stats': {
                'total_students': len(students),
                'evaluated': len([s for s in students if s['status'] == 'evaluated']),
                'pending': len([s for s in students if s['status'] != 'evaluated'])
            }
        }), 200

    except Exception as e:
        print(f"❌ Error getting results: {str(e)}")
        return jsonify({'error': str(e), 'success': False}), 500

